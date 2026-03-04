"""
Excel Generation Service using openpyxl.
Generates formatted Excel reports with charts, tables, and styled content.
"""

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, PieChart, LineChart, AreaChart, Reference
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """Generates formatted Excel reports with charts and tables."""
    
    def __init__(self):
        self.workbook = None
        self.current_row = 1
        
        # Define styles
        self.header_font = Font(bold=True, size=14)
        self.subheader_font = Font(bold=True, size=12)
        self.table_header_font = Font(bold=True, color="FFFFFF")
        self.italic_font = Font(italic=True, size=10)
        self.normal_font = Font(size=11)
        
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.right_align = Alignment(horizontal='right', vertical='center')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        self.chart_header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    def _clean_text(self, text: Any) -> str:
        """Clean and normalize text content."""
        if text is None:
            return ""
        return str(text).strip()
    
    def _add_section_header(self, worksheet, text: str, row: int) -> int:
        """Add a main section header (centered, bold)."""
        cell = worksheet.cell(row=row, column=1)
        cell.value = text
        cell.font = self.header_font
        cell.alignment = self.center_align
        worksheet.merge_cells(start_row=row, start_column=1, 
                              end_row=row, end_column=8)
        return row + 1
    
    def _add_sub_header(self, worksheet, text: str, row: int) -> int:
        """Add a sub-section header (left-aligned, bold)."""
        cell = worksheet.cell(row=row, column=1)
        cell.value = text
        cell.font = self.subheader_font
        cell.alignment = self.left_align
        return row + 1
    
    def _create_chart(self, worksheet, chart_data: Dict[str, Any], data_start_row: int, data_end_row: int, num_cols: int, chart_col: int):
        """Create and add a chart to the worksheet, positioned to the right of data."""
        chart_type = chart_data.get('type', 'bar').lower()
        title = chart_data.get('title', 'Chart')
        
        # Create appropriate chart type
        if chart_type == 'pie':
            chart = PieChart()
        elif chart_type == 'line':
            chart = LineChart()
            chart.style = 10
        elif chart_type == 'area':
            chart = AreaChart()
        else:  # Default to bar chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
        
        chart.title = title
        chart.legend.position = 'b'  # Bottom legend
        
        # Set up data references
        # Categories (labels) - first column
        categories = Reference(worksheet, min_col=1, min_row=data_start_row + 1, 
                              max_row=data_end_row)
        
        # Values - remaining columns
        for col_idx in range(2, num_cols + 1):
            values = Reference(worksheet, min_col=col_idx, min_row=data_start_row, 
                              max_row=data_end_row)
            chart.add_data(values, titles_from_data=True)
        
        chart.set_categories(categories)
        
        # Set chart size
        chart.width = 15
        chart.height = 10
        
        # Position chart to the right of the data table
        chart_cell = f"{get_column_letter(chart_col)}{data_start_row}"
        worksheet.add_chart(chart, chart_cell)
    
    def _auto_fit_columns(self, worksheet):
        """Auto-fit column widths for a worksheet."""
        for col_idx in range(1, worksheet.max_column + 1):
            max_length = 10
            column_letter = get_column_letter(col_idx)
            
            for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, min(cell_length, 50))
            
            worksheet.column_dimensions[column_letter].width = max_length + 2
    
    def generate_report(self, report_data: Dict[str, Any]) -> bytes:
        """
        Generate a complete Excel report from the provided data.
        
        Sheet 1: Key Insights and Tables
        Sheet 2: Charts and Chart Data
        
        Args:
            report_data: Dictionary containing:
                - summary_points: List of key insights
                - content: Response text (if no summary)
                - charts: List of chart data
                - tables: List of table data
        
        Returns:
            Excel file as bytes
        """
        self.workbook = Workbook()
        
        # ===== SHEET 1: Insights & Tables =====
        sheet1 = self.workbook.active
        sheet1.title = "Insights & Tables"
        current_row = 1
        
        # Main Header
        current_row = self._add_section_header(sheet1, "Asset Manager - FINANCIAL REPORT", current_row)
        
        cell = sheet1.cell(row=current_row, column=1)
        cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.font = self.italic_font
        cell.alignment = self.center_align
        sheet1.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=8)
        current_row += 2
        
        # ===== KEY INSIGHTS =====
        summary_points = report_data.get('summary_points', [])
        if summary_points:
            current_row = self._add_sub_header(sheet1, "KEY INSIGHTS", current_row)
            current_row += 1
            
            for idx, point in enumerate(summary_points, 1):
                cell = sheet1.cell(row=current_row, column=1)
                cell.value = f"{idx}. {self._clean_text(point)}"
                cell.alignment = self.left_align
                cell.font = self.normal_font
                # Merge cells across columns to display full text in single row
                sheet1.merge_cells(start_row=current_row, start_column=1,
                                   end_row=current_row, end_column=10)
                current_row += 1
            
            current_row += 2
        
        # ===== RESPONSE TEXT =====
        content = report_data.get('content', '')
        if content and not summary_points:
            current_row = self._add_sub_header(sheet1, "RESPONSE", current_row)
            current_row += 1
            
            cell = sheet1.cell(row=current_row, column=1)
            cell.value = self._clean_text(content)
            cell.alignment = self.left_align
            cell.font = self.normal_font
            sheet1.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=8)
            current_row += 3
        
        # ===== TABLES =====
        tables = report_data.get('tables', [])
        for table_idx, table in enumerate(tables):
            headers = table.get('headers', [])
            rows = table.get('full_data') or table.get('rows', [])
            
            if not headers or not rows:
                continue
            
            table_title = table.get('title', f'Table {table_idx + 1}')
            table_description = table.get('description', '')
            
            current_row = self._add_sub_header(sheet1, table_title.upper(), current_row)
            if table_description:
                cell = sheet1.cell(row=current_row, column=1)
                cell.value = self._clean_text(table_description)
                cell.font = self.italic_font
                current_row += 1
            current_row += 1
            
            # Add table headers
            for col_idx, header in enumerate(headers, 1):
                cell = sheet1.cell(row=current_row, column=col_idx)
                cell.value = self._clean_text(header).upper()
                cell.font = self.table_header_font
                cell.alignment = self.center_align
                cell.fill = self.header_fill
                cell.border = self.thin_border
            current_row += 1
            
            # Add table data rows
            for row_idx, row in enumerate(rows):
                for col_idx, cell_value in enumerate(row, 1):
                    cell = sheet1.cell(row=current_row, column=col_idx)
                    cell.value = cell_value if isinstance(cell_value, (int, float)) else self._clean_text(cell_value)
                    cell.border = self.thin_border
                    
                    # Alternate row colors
                    if row_idx % 2 == 0:
                        cell.fill = self.alt_row_fill
                    
                    # Right-align numbers
                    if isinstance(cell_value, (int, float)):
                        cell.alignment = self.right_align
                
                current_row += 1
            
            current_row += 2
        
        # Auto-fit columns for Sheet 1
        self._auto_fit_columns(sheet1)
        
        # ===== SHEET 2: Charts =====
        charts = report_data.get('charts', [])
        if charts:
            sheet2 = self.workbook.create_sheet(title="Charts & Data")
            chart_sheet_row = 1
            
            # Sheet 2 Header
            chart_sheet_row = self._add_section_header(sheet2, "CHARTS & VISUALIZATIONS", chart_sheet_row)
            
            cell = sheet2.cell(row=chart_sheet_row, column=1)
            cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            cell.font = self.italic_font
            cell.alignment = self.center_align
            sheet2.merge_cells(start_row=chart_sheet_row, start_column=1,
                               end_row=chart_sheet_row, end_column=8)
            chart_sheet_row += 2
            
            for chart_idx, chart in enumerate(charts):
                chart_data_list = chart.get('data', [])
                if isinstance(chart_data_list, dict):
                    chart_data_list = chart_data_list.get('data', [])
                
                if not chart_data_list or not isinstance(chart_data_list, list) or len(chart_data_list) == 0:
                    continue
                
                chart_title = chart.get('title', f'Chart {chart_idx + 1}')
                chart_description = chart.get('description', '')
                
                # Add chart section header
                chart_sheet_row = self._add_sub_header(sheet2, chart_title.upper(), chart_sheet_row)
                if chart_description:
                    cell = sheet2.cell(row=chart_sheet_row, column=1)
                    cell.value = self._clean_text(chart_description)
                    cell.font = self.italic_font
                    chart_sheet_row += 1
                chart_sheet_row += 1
                
                # Get keys from first data item
                keys = list(chart_data_list[0].keys())
                num_cols = len(keys)
                
                # Record data start row for chart reference
                data_start_row = chart_sheet_row
                
                # Add headers for chart data table
                for col_idx, key in enumerate(keys, 1):
                    cell = sheet2.cell(row=chart_sheet_row, column=col_idx)
                    cell.value = str(key).upper()
                    cell.font = Font(bold=True)
                    cell.alignment = self.center_align
                    cell.fill = self.chart_header_fill
                    cell.border = self.thin_border
                chart_sheet_row += 1
                
                # Add chart data rows
                for item in chart_data_list:
                    for col_idx, key in enumerate(keys, 1):
                        cell = sheet2.cell(row=chart_sheet_row, column=col_idx)
                        value = item.get(key, '')
                        cell.value = value if isinstance(value, (int, float)) else self._clean_text(value)
                        cell.border = self.thin_border
                        if isinstance(value, (int, float)):
                            cell.alignment = self.right_align
                    chart_sheet_row += 1
                
                data_end_row = chart_sheet_row - 1
                
                # Create and add chart to the right of the data table
                try:
                    # Position chart 2 columns to the right of the data
                    chart_col = num_cols + 2
                    self._create_chart(
                        sheet2, chart, data_start_row, data_end_row, num_cols, chart_col
                    )
                except Exception as e:
                    print(f"Error creating chart: {e}")
                
                # Add spacing after data table (chart is to the right, so just need row spacing)
                chart_sheet_row += 3
            
            # Auto-fit columns for Sheet 2
            self._auto_fit_columns(sheet2)
        
        # Save to bytes buffer
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def generate_single_table_export(
        self, 
        headers: List[str], 
        rows: List[List[Any]], 
        title: str = "Table",
        description: str = None
    ) -> bytes:
        """
        Generate an Excel file for a single table with formatting.
        
        Args:
            headers: List of column headers
            rows: List of data rows
            title: Table title
            description: Optional table description
        
        Returns:
            Excel file as bytes
        """
        self.workbook = Workbook()
        worksheet = self.workbook.active
        worksheet.title = title[:31] if title else "Table"  # Excel sheet name limit
        
        current_row = 1
        
        # Add title header
        current_row = self._add_section_header(worksheet, title.upper() if title else "TABLE DATA", current_row)
        
        # Add timestamp
        cell = worksheet.cell(row=current_row, column=1)
        cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.font = self.italic_font
        cell.alignment = self.center_align
        worksheet.merge_cells(start_row=current_row, start_column=1,
                              end_row=current_row, end_column=len(headers) if headers else 8)
        current_row += 1
        
        # Add description if provided
        if description:
            current_row += 1
            cell = worksheet.cell(row=current_row, column=1)
            cell.value = self._clean_text(description)
            cell.font = self.italic_font
            cell.alignment = self.left_align
            worksheet.merge_cells(start_row=current_row, start_column=1,
                                  end_row=current_row, end_column=len(headers) if headers else 8)
            current_row += 1
        
        current_row += 1  # Add spacing
        
        # Add table headers with formatting
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col_idx)
            cell.value = self._clean_text(header).upper()
            cell.font = self.table_header_font
            cell.alignment = self.center_align
            cell.fill = self.header_fill
            cell.border = self.thin_border
        current_row += 1
        
        # Add data rows with formatting
        for row_idx, row in enumerate(rows):
            for col_idx, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=current_row, column=col_idx)
                cell.value = cell_value if isinstance(cell_value, (int, float)) else self._clean_text(cell_value)
                cell.border = self.thin_border
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = self.alt_row_fill
                
                # Right-align numbers
                if isinstance(cell_value, (int, float)):
                    cell.alignment = self.right_align
            
            current_row += 1
        
        # Auto-fit columns
        self._auto_fit_columns(worksheet)
        
        # Save to bytes buffer
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()


# Singleton instance
excel_service = ExcelReportGenerator()

